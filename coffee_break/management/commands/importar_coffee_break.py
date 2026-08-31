"""Importa o histórico da planilha "CONTROLE COFFE ASCOM.xlsx".

Uso:
    python manage.py importar_coffee_break caminho/planilha.xlsx --usuario tiago
    python manage.py importar_coffee_break caminho/planilha.xlsx --usuario tiago --dry-run

O comando é transacional e idempotente: rodar de novo atualiza os registros
existentes sem duplicar fornecedores, contratos, lotes ou solicitações.

Particularidades da planilha tratadas aqui:
- números de solicitação que o Excel converteu em datas (ex.: "02/2026"
  virou 01/02/2026 com formato mm/yyyy) voltam a ser texto;
- datas de evento em formato livre ("23, 24 e 25/03", "12 à 15/05") são
  preservadas como texto e, quando possível, também estruturadas;
- a aba "Controle de Ordem Bancária" alimenta os marcos de OB das
  solicitações de 2026;
- linhas com "CANCELADO" (ou quantidade zero) entram como canceladas, fora
  do consumo do lote, sem perder o histórico.
"""

import datetime as dt
import re
import unicodedata

from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from cadastros.models import Municipio
from coffee_break.models import (
    ContratoCoffeeBreak,
    Fornecedor,
    LoteCoffeeBreak,
    SolicitacaoCoffeeBreak,
)

FISCAL_RESPONSAVEL = "Janine Lacerda do Prado"
OBJETO_CONTRATO = "Aquisição/Contratação de Serviço de Coffee Break — PREG-e 494/2023"

FORNECEDORES = {
    "FAVO": {
        "razao_social": "PADARIA E CONFEITARIA FAVO E MEL LTDA",
        "cnpj": "35014719000166",
        "contato": "Cesar Fernandes",
        "telefone": "41 9979-2659",
        "email": "contato@favoemel.com.br",
    },
    "GIACOMINI": {
        "razao_social": "GIACOMINI E CARVALHO LTDA",
        "cnpj": "45549407000100",
        "contato": "Felipe Giacomini",
        "telefone": "45 9831-2335",
        "email": "compras.giacominiecarvalho@gmail.com",
    },
    "PANE": {
        "razao_social": "PANE PERFETTO SOLUCOES ALIMENTICIAS LTDA",
        # A planilha repete o CNPJ da Giacomini na aba do lote 4 — erro de
        # origem; fica em branco até confirmação.
        "cnpj": "",
        "contato": "Pane Perfetto",
        "telefone": "41 9116-4715",
        "email": "paneperfetto@outlook.com",
    },
}

CONTRATOS = {
    "0762/2024": {"fornecedor": "FAVO", "numero_gms": "7339/2024", "termo_aditivo": "0355/2025"},
    "0923/2024": {"fornecedor": "GIACOMINI", "numero_gms": "8811/2024", "termo_aditivo": "0001/2025"},
    "0823/2024": {"fornecedor": "GIACOMINI", "numero_gms": "7749/2024", "termo_aditivo": "0356/2025"},
    "0150/2025": {"fornecedor": "PANE", "numero_gms": "1364/2025", "termo_aditivo": ""},
    "0130/2025": {"fornecedor": "GIACOMINI", "numero_gms": "1182/2025", "termo_aditivo": ""},
}

# Aba -> lote. As abas ocultas são a vigência anterior (2024/2025); as
# "- 2026" são o exercício corrente, com empenho próprio.
ABAS_LOTES = {
    "LOTE 1": {"contrato": "0762/2024", "numero": 1, "exercicio": "2025", "empenho": ""},
    " LOTE 1- 2026": {"contrato": "0762/2024", "numero": 1, "exercicio": "2026", "empenho": "2026NE030208"},
    "LOTE 2": {"contrato": "0923/2024", "numero": 2, "exercicio": "2025", "empenho": "2025NE009466"},
    "LOTE 2- 2026": {"contrato": "0923/2024", "numero": 2, "exercicio": "2026", "empenho": "2025NE135230"},
    "LOTE 3": {"contrato": "0823/2024", "numero": 3, "exercicio": "2025", "empenho": "2025NE007298"},
    "LOTE 3- 2026": {"contrato": "0823/2024", "numero": 3, "exercicio": "2026", "empenho": "2025NE135231"},
    "LOTE 4": {"contrato": "0150/2025", "numero": 4, "exercicio": "2025", "empenho": "2025NR000093"},
    "LOTE 5": {"contrato": "0130/2025", "numero": 5, "exercicio": "2025", "empenho": "2025NE135234"},
    "LOTE 5- 2026": {"contrato": "0130/2025", "numero": 5, "exercicio": "2026", "empenho": "2025NE135234"},
}

ABA_ORDEM_BANCARIA = "Controle de Ordem Bancária"
ABA_GERAL = "Geral"

VAZIOS = {"", "-", "--", "—", "?"}


def _limpa(texto):
    return re.sub(r"\s+", " ", str(texto or "").strip())


def _norm(texto):
    texto = _limpa(texto).lower()
    return "".join(
        c
        for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )


def _vazio(valor):
    return valor is None or _limpa(valor) in VAZIOS


def _como_data(valor):
    """Datas reais da planilha (datetime/date) -> date; resto -> None."""
    if isinstance(valor, dt.datetime):
        return valor.date()
    if isinstance(valor, dt.date):
        return valor
    return None


def _parse_data_texto(texto):
    m = re.match(r"^(\d{1,2})/(\d{1,2})/(\d{4})$", _limpa(texto))
    if not m:
        return None
    try:
        return dt.date(int(m.group(3)), int(m.group(2)), int(m.group(1)))
    except ValueError:
        return None


def identificador_como_texto(valor):
    """Identificadores institucionais sempre como texto.

    O Excel converteu "02/2026" (número da solicitação) em 01/02/2026;
    aqui a data mensal volta a ser "MM/AAAA". Números viram string sem
    o ".0" do float.
    """
    if valor is None:
        return ""
    if isinstance(valor, (dt.datetime, dt.date)):
        return f"{valor.month:02d}/{valor.year}"
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    if isinstance(valor, int):
        return str(valor)
    texto = _limpa(valor)
    return "" if texto in VAZIOS else texto


def _como_quantidade(valor):
    if isinstance(valor, (int, float)):
        return int(valor)
    m = re.match(r"^(\d+)", _limpa(valor))
    return int(m.group(1)) if m else None


def parse_periodo_livre(texto, ano_padrao):
    """"23, 24 e 25/03" ou "12 à 15/05" -> (inicio, fim); None se ambíguo."""
    bruto = _limpa(texto)
    if not bruto:
        return None, None
    ano = ano_padrao
    m_ano = re.search(r"/(\d{4})", bruto)
    if m_ano:
        ano = int(m_ano.group(1))
    partes = re.findall(r"(\d{1,2})(?:/(\d{1,2}))?", re.sub(r"/\d{4}", "", bruto))
    if not partes:
        return None, None
    meses = [int(m) if m else None for _, m in partes]
    ultimo = None
    for i in range(len(meses) - 1, -1, -1):
        if meses[i] is not None:
            ultimo = meses[i]
        elif ultimo is not None:
            meses[i] = ultimo
    datas = []
    for (dia, _), mes in zip(partes, meses):
        if not mes:
            return None, None
        try:
            datas.append(dt.date(ano, mes, int(dia)))
        except ValueError:
            return None, None
    return min(datas), max(datas)


class Command(BaseCommand):
    help = "Importa o controle de coffee break da ASCOM (planilha .xlsx)."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", help="Caminho da planilha .xlsx")
        parser.add_argument(
            "--usuario", required=True, help="Username do criador dos registros"
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Simula a importação sem gravar nada no banco",
        )

    def handle(self, *args, **opts):
        try:
            import openpyxl
        except ImportError:  # pragma: no cover
            raise CommandError("openpyxl não está instalado (pip install openpyxl).")

        User = get_user_model()
        try:
            usuario = User.objects.get(username=opts["usuario"])
        except User.DoesNotExist:
            raise CommandError(f"Usuário '{opts['usuario']}' não existe.")

        try:
            wb = openpyxl.load_workbook(opts["xlsx_path"], data_only=True)
        except FileNotFoundError:
            raise CommandError(f"Planilha não encontrada: {opts['xlsx_path']}")

        self.resumo = {"criados": 0, "atualizados": 0, "ignorados": 0}
        self.avisos = []

        with transaction.atomic():
            fornecedores = self._importar_fornecedores()
            contratos = self._importar_contratos(fornecedores)
            lotes = {}
            for nome_aba, config in ABAS_LOTES.items():
                if nome_aba not in wb.sheetnames:
                    self.avisos.append(f"Aba '{nome_aba}' não encontrada — pulada.")
                    continue
                lote = self._importar_lote(wb, nome_aba, config, contratos)
                lotes[(config["numero"], config["exercicio"])] = lote
                self._importar_solicitacoes(wb[nome_aba], lote, usuario)
            if ABA_ORDEM_BANCARIA in wb.sheetnames:
                self._importar_ordem_bancaria(wb[ABA_ORDEM_BANCARIA], lotes)
            else:
                self.avisos.append(
                    f"Aba '{ABA_ORDEM_BANCARIA}' não encontrada — marcos de OB não importados."
                )
            if opts["dry_run"]:
                transaction.set_rollback(True)

        estilo = self.style.WARNING if opts["dry_run"] else self.style.SUCCESS
        prefixo = "[dry-run] " if opts["dry_run"] else ""
        self.stdout.write(
            estilo(
                f"{prefixo}{self.resumo['criados']} criados, "
                f"{self.resumo['atualizados']} atualizados, "
                f"{self.resumo['ignorados']} ignorados, "
                f"{len(self.avisos)} inconsistências."
            )
        )
        for aviso in self.avisos:
            self.stdout.write(self.style.WARNING(f"  - {aviso}"))

    # ------------------------------------------------------------------
    # Fornecedores, contratos e lotes
    # ------------------------------------------------------------------

    def _contabilizar(self, criado):
        self.resumo["criados" if criado else "atualizados"] += 1

    def _importar_fornecedores(self):
        fornecedores = {}
        for chave, dados in FORNECEDORES.items():
            fornecedor, criado = Fornecedor.objects.update_or_create(
                razao_social=dados["razao_social"],
                defaults={
                    "cnpj": dados["cnpj"],
                    "contato": dados["contato"],
                    "telefone": dados["telefone"],
                    "email": dados["email"],
                },
            )
            self._contabilizar(criado)
            fornecedores[chave] = fornecedor
        self.avisos.append(
            "Lote 4: a planilha traz o CNPJ da Giacomini para a Pane Perfetto — "
            "CNPJ deixado em branco até confirmação."
        )
        return fornecedores

    def _importar_contratos(self, fornecedores):
        contratos = {}
        for numero, dados in CONTRATOS.items():
            contrato, criado = ContratoCoffeeBreak.objects.update_or_create(
                numero=numero,
                defaults={
                    "fornecedor": fornecedores[dados["fornecedor"]],
                    "numero_gms": dados["numero_gms"],
                    "termo_aditivo": dados["termo_aditivo"],
                    "fiscal_responsavel": FISCAL_RESPONSAVEL,
                    "objeto": OBJETO_CONTRATO,
                },
            )
            self._contabilizar(criado)
            contratos[numero] = contrato
        return contratos

    def _linha_cabecalho(self, ws):
        for linha in range(1, 8):
            valor = _limpa(ws.cell(row=linha, column=1).value)
            if _norm(valor).startswith("data da solicita"):
                return linha
        return None

    def _total_do_lote(self, ws):
        """Valor numérico logo abaixo do rótulo TOTAL na coluna J."""
        for linha in range(1, 8):
            valor = ws.cell(row=linha, column=10).value
            if isinstance(valor, (int, float)):
                return int(valor)
        return None

    def _textos_coluna(self, ws, coluna, ate_linha=10):
        textos = []
        for linha in range(1, ate_linha + 1):
            valor = ws.cell(row=linha, column=coluna).value
            if isinstance(valor, str) and valor.strip():
                textos.append(valor.strip())
        return textos

    ROTULOS_INFORMATIVOS = {
        "informacoes sobre o lote",
        "orientacoes",
        "especificacoes tecnicas gerais",
        "gastos",
        "restantes",
        "total",
    }

    def _info_lote(self, ws, geral):
        """Separa, na coluna L, cidades / objeto / orientações / observações."""
        cidades = ""
        orientacoes = []
        observacoes = []
        for texto in self._textos_coluna(ws, coluna=12):
            chave = _norm(texto)
            if chave in self.ROTULOS_INFORMATIVOS or chave.startswith("fornecedor lote"):
                continue
            if "preg-e" in chave:
                continue  # objeto do contrato, já registrado no contrato
            if not cidades:
                cidades = texto
            elif chave.startswith("obs"):
                observacoes.append(texto)
            else:
                orientacoes.append(texto)
        especificacoes = ""
        for texto in self._textos_coluna(ws, coluna=14):
            if len(texto) > 100:
                especificacoes = texto
                break
        return {
            "cidades": cidades,
            "orientacoes": "\n\n".join(orientacoes) or geral["orientacoes"],
            "especificacoes": especificacoes or geral["especificacoes"],
            "observacoes": "\n\n".join(observacoes),
        }

    def _dados_gerais(self, wb):
        orientacoes = ""
        especificacoes = ""
        if ABA_GERAL in wb.sheetnames:
            ws = wb[ABA_GERAL]
            orientacoes = str(ws["F2"].value or "").strip()
            especificacoes = str(ws["J3"].value or "").strip()
        return {"orientacoes": orientacoes, "especificacoes": especificacoes}

    def _importar_lote(self, wb, nome_aba, config, contratos):
        ws = wb[nome_aba]
        geral = getattr(self, "_geral_cache", None)
        if geral is None:
            geral = self._geral_cache = self._dados_gerais(wb)

        total = self._total_do_lote(ws)
        if total is None:
            self.avisos.append(
                f"Aba '{nome_aba}': quantitativo total não encontrado — lote pulado."
            )
            return None
        info = self._info_lote(ws, geral)

        lote, criado = LoteCoffeeBreak.objects.update_or_create(
            contrato=contratos[config["contrato"]],
            numero=config["numero"],
            exercicio=config["exercicio"],
            defaults={
                "quantidade_total": total,
                "empenho": config["empenho"],
                "municipios_texto": info["cidades"],
                "orientacoes": info["orientacoes"],
                "especificacoes_tecnicas": info["especificacoes"],
                "observacoes": info["observacoes"],
                # A vigência anterior fica inativa; só o exercício corrente
                # recebe novas solicitações.
                "ativo": config["exercicio"] == "2026",
            },
        )
        self._contabilizar(criado)
        self._vincular_municipios(lote, info["cidades"], nome_aba)
        return lote

    def _vincular_municipios(self, lote, texto_cidades, nome_aba):
        texto = re.sub(r"(?i)^cidades abrangentes:\s*", "", texto_cidades or "")
        nomes = [n.strip(" .") for n in texto.split(",") if n.strip(" .")]
        encontrados = []
        for nome in nomes:
            municipio = Municipio.objects.filter(nome__iexact=nome).first()
            if municipio:
                encontrados.append(municipio)
                continue
            # O último item costuma vir como par: "Matinhos e Pinhais".
            partes = [p.strip(" .") for p in re.split(r"\s+e\s+", nome) if p.strip(" .")]
            achados = [
                Municipio.objects.filter(nome__iexact=parte).first()
                for parte in partes
            ]
            if len(partes) > 1 and all(achados):
                encontrados.extend(achados)
            else:
                self.avisos.append(
                    f"Aba '{nome_aba}': município '{nome}' não reconhecido — "
                    "mantido apenas no texto original."
                )
        lote.municipios.set(encontrados)

    # ------------------------------------------------------------------
    # Solicitações
    # ------------------------------------------------------------------

    def _importar_solicitacoes(self, ws, lote, usuario):
        if lote is None:
            return
        cabecalho = self._linha_cabecalho(ws)
        if cabecalho is None:
            self.avisos.append(
                f"Aba '{ws.title}': cabeçalho de solicitações não encontrado."
            )
            return
        vazias_seguidas = 0
        for linha in range(cabecalho + 1, ws.max_row + 1):
            celulas = [ws.cell(row=linha, column=c).value for c in range(1, 9)]
            if all(v is None or str(v).strip() == "" for v in celulas):
                vazias_seguidas += 1
                if vazias_seguidas >= 5:
                    break
                continue
            vazias_seguidas = 0
            self._importar_linha(ws.title, linha, celulas, lote, usuario)

    def _importar_linha(self, aba, linha, celulas, lote, usuario):
        (
            data_solic, data_evento, numero, descricao, nota_fiscal,
            quantidade, protocolo, atesto,
        ) = celulas

        descricao_texto = _limpa(descricao)
        if not descricao_texto:
            self.resumo["ignorados"] += 1
            self.avisos.append(f"Aba '{aba}', linha {linha}: sem descrição — ignorada.")
            return

        observacoes = []
        data_solicitacao = _como_data(data_solic)
        if data_solicitacao is None:
            data_solicitacao = _parse_data_texto(data_solic)
        if data_solicitacao is None:
            self.resumo["ignorados"] += 1
            self.avisos.append(
                f"Aba '{aba}', linha {linha}: data da solicitação "
                f"'{_limpa(data_solic)}' inválida — linha ignorada."
            )
            return

        inicio = fim = None
        periodo_texto = ""
        data_estruturada = _como_data(data_evento)
        if data_estruturada:
            inicio = fim = data_estruturada
        elif not _vazio(data_evento):
            periodo_texto = _limpa(data_evento)[:120]
            inicio, fim = parse_periodo_livre(periodo_texto, data_solicitacao.year)
            if inicio is None:
                self.avisos.append(
                    f"Aba '{aba}', linha {linha}: período do evento "
                    f"'{periodo_texto}' mantido apenas como texto."
                )

        numero_texto = identificador_como_texto(numero)
        nf_texto = identificador_como_texto(nota_fiscal)
        protocolo_texto = identificador_como_texto(protocolo)

        quantidade_valor = _como_quantidade(quantidade)
        if quantidade_valor is None:
            quantidade_valor = 0
            self.avisos.append(
                f"Aba '{aba}', linha {linha}: quantidade '{_limpa(quantidade)}' "
                "inválida — registrada como 0 (cancelada)."
            )

        campos_texto = f"{descricao_texto} {nf_texto} {protocolo_texto}".upper()
        cancelada = "CANCELADO" in campos_texto or quantidade_valor == 0
        if cancelada:
            nf_texto = "" if "CANCELADO" in nf_texto.upper() else nf_texto
            protocolo_texto = (
                "" if "CANCELADO" in protocolo_texto.upper() else protocolo_texto
            )

        data_atesto = _como_data(atesto)
        if data_atesto is None and not _vazio(atesto) and not (
            isinstance(atesto, str) and "CANCELADO" in atesto.upper()
        ):
            observacoes.append(
                f"Data de atesto na planilha: '{_limpa(atesto)}' (não interpretada)."
            )

        defaults = {
            "data_solicitacao": data_solicitacao,
            "data_inicio_evento": inicio,
            "data_fim_evento": fim,
            "periodo_evento_texto": periodo_texto,
            "descricao_evento": descricao_texto,
            "quantidade": quantidade_valor,
            "numero_nota_fiscal": nf_texto[:30],
            "protocolo_pagamento": protocolo_texto[:30],
            "data_atesto_gaf": data_atesto,
            "cancelada": cancelada,
            "observacoes": "\n".join(observacoes),
        }
        if cancelada:
            defaults["motivo_cancelamento"] = "Registro cancelado na planilha de origem."

        chave = {"lote": lote}
        if numero_texto:
            chave["numero"] = numero_texto
        else:
            chave["data_solicitacao"] = data_solicitacao
            chave["descricao_evento"] = descricao_texto
            defaults.pop("data_solicitacao")
            defaults.pop("descricao_evento")

        solicitacao, criada = SolicitacaoCoffeeBreak.objects.update_or_create(
            **chave,
            defaults=defaults,
            create_defaults={
                **defaults,
                "criado_por": usuario,
                "numero": numero_texto,
            },
        )
        self._contabilizar(criada)

    # ------------------------------------------------------------------
    # Controle de ordem bancária
    # ------------------------------------------------------------------

    def _importar_ordem_bancaria(self, ws, lotes):
        # Dois grupos de colunas lado a lado: A-F e H-M.
        for coluna_inicial in (1, 8):
            lote_atual = None
            for linha in range(1, ws.max_row + 1):
                primeira = ws.cell(row=linha, column=coluna_inicial).value
                if isinstance(primeira, str):
                    m = re.match(r"\s*LOTE\s*(\d)", primeira)
                    if m and "-" in primeira:
                        numero = int(m.group(1))
                        # A aba de OB acompanha o exercício corrente (2026).
                        lote_atual = lotes.get((numero, "2026"))
                        if lote_atual is None:
                            self.avisos.append(
                                f"Ordem bancária: lote {numero}/2026 não importado — "
                                "bloco ignorado."
                            )
                        continue
                    if _norm(primeira).startswith("n° da solicita") or _norm(
                        primeira
                    ).startswith("no da solicita"):
                        continue
                if lote_atual is None:
                    continue
                numero_texto = identificador_como_texto(primeira)
                if not numero_texto:
                    continue
                descricao = _limpa(ws.cell(row=linha, column=coluna_inicial + 1).value)
                ob_emitida = _como_data(ws.cell(row=linha, column=coluna_inicial + 4).value)
                envio = _como_data(ws.cell(row=linha, column=coluna_inicial + 5).value)
                solicitacao = SolicitacaoCoffeeBreak.objects.filter(
                    lote=lote_atual, numero=numero_texto
                ).first()
                if solicitacao is None:
                    self.avisos.append(
                        f"Ordem bancária: solicitação {numero_texto} "
                        f"({descricao[:40]}...) não encontrada no "
                        f"{lote_atual.rotulo_curto}."
                    )
                    continue
                mudou = False
                if ob_emitida and solicitacao.data_ordem_bancaria != ob_emitida:
                    solicitacao.data_ordem_bancaria = ob_emitida
                    mudou = True
                if envio and solicitacao.data_envio_empresa != envio:
                    solicitacao.data_envio_empresa = envio
                    mudou = True
                if mudou:
                    solicitacao.save(
                        update_fields=[
                            "data_ordem_bancaria",
                            "data_envio_empresa",
                            "atualizado_em",
                        ]
                    )
                    self.resumo["atualizados"] += 1
