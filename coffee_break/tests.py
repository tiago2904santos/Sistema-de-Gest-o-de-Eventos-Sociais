import datetime as dt
import io

from django.core.exceptions import ValidationError
from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.management import call_command
from django.test import TestCase
from django.urls import reverse

from accounts.models import Modulo, Setor
from accounts.modulos import codigos_modulos_do_usuario, usuario_tem_modulo
from cadastros.models import Estado, Municipio, Regiao
from solicitacoes.permissions import GRUPO_ADMINISTRADOR

from .forms import SolicitacaoCoffeeBreakForm
from .management.commands.importar_coffee_break import (
    identificador_como_texto,
    parse_periodo_livre,
)
from .models import (
    ContratoCoffeeBreak,
    Fornecedor,
    LoteCoffeeBreak,
    SituacaoFinanceira,
    SolicitacaoCoffeeBreak,
    normalizar_cnpj,
)
from .permissions import CODIGO_MODULO
from . import services

User = get_user_model()


class BaseCoffeeBreakTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.setor_ascom = Setor.objects.get(nome="ASCOM")
        cls.modulo = Modulo.objects.get(codigo=CODIGO_MODULO)

        cls.ascom = User.objects.create_user("ascom", password="x")
        cls.ascom.setores.add(cls.setor_ascom)
        cls.sem_modulo = User.objects.create_user("comum", password="x")
        cls.superusuario = User.objects.create_superuser("root", password="x")
        cls.admin_modulo = User.objects.create_user("admin-coffee", password="x")
        cls.admin_modulo.setores.add(cls.setor_ascom)
        grupo_admin, _ = Group.objects.get_or_create(name=GRUPO_ADMINISTRADOR)
        cls.admin_modulo.groups.add(grupo_admin)

        cls.fornecedor = Fornecedor.objects.create(
            razao_social="PADARIA E CONFEITARIA FAVO E MEL LTDA",
            cnpj="35014719000166",
            email="contato@favoemel.com.br",
        )
        cls.contrato = ContratoCoffeeBreak.objects.create(
            fornecedor=cls.fornecedor,
            numero="0762/2024",
            numero_gms="7339/2024",
            fiscal_responsavel="Janine Lacerda do Prado",
        )
        cls.lote = LoteCoffeeBreak.objects.create(
            contrato=cls.contrato,
            numero=1,
            exercicio="2026",
            quantidade_total=100,
            empenho="2026NE030208",
        )

    def criar_solicitacao(self, **kwargs):
        dados = {
            "lote": self.lote,
            "data_solicitacao": dt.date(2026, 8, 1),
            "descricao_evento": "Evento de teste",
            "quantidade": 30,
            "criado_por": self.ascom,
        }
        dados.update(kwargs)
        return SolicitacaoCoffeeBreak.objects.create(**dados)


# ---------------------------------------------------------------------------
# Autorização por módulo
# ---------------------------------------------------------------------------

class AutorizacaoModuloTests(BaseCoffeeBreakTestCase):
    ROTAS = [
        ("coffee_break:painel", []),
        ("coffee_break:lotes", []),
        ("coffee_break:solicitacoes", []),
        ("coffee_break:nova", []),
    ]

    def test_seed_criou_modulo_e_setor(self):
        self.assertTrue(Modulo.objects.filter(codigo=CODIGO_MODULO).exists())
        self.assertIn(self.setor_ascom, self.modulo.setores.all())

    def test_usuario_ascom_acessa_todas_as_telas(self):
        self.client.force_login(self.ascom)
        for rota, args in self.ROTAS:
            resposta = self.client.get(reverse(rota, args=args))
            self.assertEqual(resposta.status_code, 200, rota)

    def test_usuario_sem_modulo_recebe_403_por_url_direta(self):
        self.client.force_login(self.sem_modulo)
        solicitacao = self.criar_solicitacao()
        rotas = [reverse(rota, args=args) for rota, args in self.ROTAS]
        rotas += [
            reverse("coffee_break:detalhe", args=[solicitacao.pk]),
            reverse("coffee_break:editar", args=[solicitacao.pk]),
            reverse("coffee_break:lote_detalhe", args=[self.lote.pk]),
        ]
        for url in rotas:
            self.assertEqual(self.client.get(url).status_code, 403, url)
        # Ações POST também são bloqueadas no backend.
        self.assertEqual(
            self.client.post(
                reverse("coffee_break:cancelar", args=[solicitacao.pk])
            ).status_code,
            403,
        )

    def test_anonimo_redireciona_para_login(self):
        resposta = self.client.get(reverse("coffee_break:painel"))
        self.assertEqual(resposta.status_code, 302)
        self.assertIn(reverse("accounts:login"), resposta.url)

    def test_superusuario_tem_visao_global(self):
        self.client.force_login(self.superusuario)
        self.assertEqual(
            self.client.get(reverse("coffee_break:painel")).status_code, 200
        )
        self.assertTrue(usuario_tem_modulo(self.superusuario, CODIGO_MODULO))

    def test_modulo_inativo_bloqueia_ascom(self):
        Modulo.objects.filter(pk=self.modulo.pk).update(ativo=False)
        self.assertFalse(usuario_tem_modulo(self.ascom, CODIGO_MODULO))

    def test_portal_mostra_modulo_conforme_acesso(self):
        """O hub lista o módulo só para quem tem o código ASCOM_COFFEE_BREAK."""
        self.client.force_login(self.ascom)
        resposta = self.client.get(reverse("core:home"))
        self.assertContains(resposta, reverse("coffee_break:painel"))

        self.client.force_login(self.sem_modulo)
        resposta = self.client.get(reverse("core:home"))
        self.assertNotContains(resposta, reverse("coffee_break:painel"))

    def test_navbar_contextual_dentro_e_fora_do_modulo(self):
        """Dentro do módulo só aparecem itens dele; fora, ele não aparece."""
        self.client.force_login(self.ascom)
        # No dashboard de Eventos Sociais não há navegação do Coffee Break.
        resposta = self.client.get(reverse("dashboard:index"))
        self.assertNotContains(
            resposta, f'href="{reverse("coffee_break:painel")}"'
        )
        self.assertContains(resposta, f'href="{reverse("solicitacoes:lista")}"')
        # Dentro do Coffee Break só há navegação do próprio módulo.
        resposta = self.client.get(reverse("coffee_break:painel"))
        self.assertContains(resposta, f'href="{reverse("coffee_break:lotes")}"')
        self.assertNotContains(
            resposta, f'href="{reverse("solicitacoes:lista")}"'
        )
        self.assertContains(resposta, f'href="{reverse("core:home")}"')

    def test_codigos_modulos_do_usuario(self):
        self.assertIn(CODIGO_MODULO, codigos_modulos_do_usuario(self.ascom))
        self.assertEqual(codigos_modulos_do_usuario(self.sem_modulo), set())
        self.assertIn(CODIGO_MODULO, codigos_modulos_do_usuario(self.superusuario))


# ---------------------------------------------------------------------------
# Modelos: fornecedor, contrato, lote e cálculos
# ---------------------------------------------------------------------------

class ModelosTests(BaseCoffeeBreakTestCase):
    def test_criacao_de_fornecedor_normaliza_cnpj(self):
        fornecedor = Fornecedor.objects.create(
            razao_social="GIACOMINI E CARVALHO LTDA",
            cnpj="45.549.407/0001-00",
        )
        self.assertEqual(fornecedor.cnpj, "45549407000100")
        self.assertEqual(fornecedor.cnpj_formatado, "45.549.407/0001-00")

    def test_cnpj_invalido_rejeitado_na_validacao(self):
        fornecedor = Fornecedor(razao_social="X LTDA", cnpj="123")
        with self.assertRaises(ValidationError):
            fornecedor.full_clean()

    def test_normalizar_cnpj(self):
        self.assertEqual(normalizar_cnpj("45.549.407/0001-00"), "45549407000100")
        self.assertEqual(normalizar_cnpj(None), "")

    def test_criacao_de_contrato_e_lote(self):
        self.assertEqual(str(self.contrato), "Contrato 0762/2024 — PADARIA E CONFEITARIA FAVO E MEL LTDA")
        self.assertEqual(self.lote.rotulo_curto, "Lote 1 (2026)")

    def test_consumido_e_restante_calculados(self):
        self.criar_solicitacao(quantidade=30)
        self.criar_solicitacao(quantidade=20, numero="02/2026")
        self.assertEqual(self.lote.quantidade_consumida, 50)
        self.assertEqual(self.lote.saldo_restante, 50)
        anotado = LoteCoffeeBreak.objects.com_consumo().get(pk=self.lote.pk)
        self.assertEqual(anotado.consumido, 50)
        self.assertEqual(anotado.restante, 50)

    def test_cancelada_nao_conta_no_consumo(self):
        self.criar_solicitacao(quantidade=30)
        self.criar_solicitacao(quantidade=60, cancelada=True, numero="02/2026")
        self.assertEqual(self.lote.quantidade_consumida, 30)

    def test_lotes_em_alerta(self):
        self.criar_solicitacao(quantidade=90)  # restam 10 de 100 (10% <= 15%)
        anotados = list(LoteCoffeeBreak.objects.com_consumo())
        em_alerta = services.lotes_em_alerta(anotados)
        self.assertEqual([lote.pk for lote in em_alerta], [self.lote.pk])


# ---------------------------------------------------------------------------
# Saldo e validações
# ---------------------------------------------------------------------------

class SaldoTests(BaseCoffeeBreakTestCase):
    def test_bloqueia_quantidade_acima_do_saldo(self):
        self.criar_solicitacao(quantidade=80)
        nova = SolicitacaoCoffeeBreak(
            lote=self.lote,
            descricao_evento="Estouro",
            quantidade=21,
            criado_por=self.ascom,
        )
        with self.assertRaises(ValidationError):
            services.salvar_com_saldo(nova)
        self.assertEqual(self.lote.solicitacoes.count(), 1)

    def test_edicao_nao_conta_a_propria_quantidade(self):
        solicitacao = self.criar_solicitacao(quantidade=80)
        solicitacao.quantidade = 100
        services.salvar_com_saldo(solicitacao)  # não deve levantar erro
        self.assertEqual(self.lote.quantidade_consumida, 100)

    def test_quantidade_zero_e_negativa_rejeitadas_no_form(self):
        for quantidade in ("0", "-5"):
            form = SolicitacaoCoffeeBreakForm(
                data={
                    "lote": self.lote.pk,
                    "data_solicitacao": "2026-08-01",
                    "descricao_evento": "Teste",
                    "quantidade": quantidade,
                }
            )
            self.assertFalse(form.is_valid(), quantidade)
            self.assertIn("quantidade", form.errors)

    def test_periodo_invalido_rejeitado(self):
        form = SolicitacaoCoffeeBreakForm(
            data={
                "lote": self.lote.pk,
                "data_solicitacao": "2026-08-01",
                "descricao_evento": "Teste",
                "quantidade": "10",
                "data_inicio_evento": "2026-08-10",
                "data_fim_evento": "2026-08-05",
            }
        )
        self.assertFalse(form.is_valid())
        self.assertIn("data_fim_evento", form.errors)

    def test_marcos_financeiros_exigem_sequencia(self):
        form = SolicitacaoCoffeeBreakForm(
            data={
                "lote": self.lote.pk,
                "data_solicitacao": "2026-08-01",
                "descricao_evento": "Teste",
                "quantidade": "10",
                "data_ordem_bancaria": "2026-08-10",
            }
        )
        self.assertFalse(form.is_valid())
        self.assertIn("data_ordem_bancaria", form.errors)

    def test_periodo_textual_e_datas_nao_sao_aceitos_juntos_em_novo_registro(self):
        form = SolicitacaoCoffeeBreakForm(
            data={
                "lote": self.lote.pk,
                "data_solicitacao": "2026-08-01",
                "descricao_evento": "Teste",
                "quantidade": "10",
                "data_inicio_evento": "2026-08-10",
                "data_fim_evento": "2026-08-11",
                "periodo_evento_texto": "10 e 11/08",
            }
        )
        self.assertFalse(form.is_valid())
        self.assertIn("periodo_evento_texto", form.errors)

    def test_cancelamento_devolve_saldo_e_registra_auditoria(self):
        solicitacao = self.criar_solicitacao(quantidade=80)
        services.cancelar(solicitacao, self.ascom, motivo="Evento adiado")
        solicitacao.refresh_from_db()
        self.assertTrue(solicitacao.cancelada)
        self.assertEqual(solicitacao.cancelada_por, self.ascom)
        self.assertEqual(solicitacao.motivo_cancelamento, "Evento adiado")
        self.assertIsNotNone(solicitacao.cancelada_em)
        self.assertEqual(self.lote.saldo_restante, 100)

    def test_reativacao_revalida_saldo(self):
        cancelada = self.criar_solicitacao(quantidade=80, cancelada=True)
        self.criar_solicitacao(quantidade=50, numero="02/2026")
        with self.assertRaises(ValidationError):
            services.reativar(cancelada)
        cancelada.refresh_from_db()
        self.assertTrue(cancelada.cancelada)


# ---------------------------------------------------------------------------
# Situação financeira derivada
# ---------------------------------------------------------------------------

class SituacaoFinanceiraTests(BaseCoffeeBreakTestCase):
    def test_derivacao_por_marcos(self):
        s = self.criar_solicitacao()
        self.assertEqual(
            s.situacao_financeira, SituacaoFinanceira.AGUARDANDO_NOTA_FISCAL
        )
        s.numero_nota_fiscal = "8046"
        self.assertEqual(
            s.situacao_financeira, SituacaoFinanceira.AGUARDANDO_PROTOCOLO
        )
        s.protocolo_pagamento = "25.419.856-0"
        self.assertEqual(s.situacao_financeira, SituacaoFinanceira.AGUARDANDO_ATESTO)
        s.data_atesto_gaf = dt.date(2026, 2, 18)
        self.assertEqual(
            s.situacao_financeira, SituacaoFinanceira.AGUARDANDO_ORDEM_BANCARIA
        )
        s.data_ordem_bancaria = dt.date(2026, 2, 24)
        self.assertEqual(
            s.situacao_financeira, SituacaoFinanceira.AGUARDANDO_ENVIO_EMPRESA
        )
        s.data_envio_empresa = dt.date(2026, 6, 16)
        self.assertEqual(s.situacao_financeira, SituacaoFinanceira.CONCLUIDA)

    def test_cancelada_prevalece(self):
        s = self.criar_solicitacao(
            cancelada=True,
            numero_nota_fiscal="8046",
            protocolo_pagamento="25.419.856-0",
            data_atesto_gaf=dt.date(2026, 2, 18),
            data_ordem_bancaria=dt.date(2026, 2, 24),
            data_envio_empresa=dt.date(2026, 6, 16),
        )
        self.assertEqual(s.situacao_financeira, SituacaoFinanceira.CANCELADA)


# ---------------------------------------------------------------------------
# Views: criação, edição, filtros, busca, ordenação e paginação
# ---------------------------------------------------------------------------

class ViewsTests(BaseCoffeeBreakTestCase):
    def setUp(self):
        self.client.force_login(self.ascom)

    def dados_post(self, **kwargs):
        dados = {
            "lote": self.lote.pk,
            "data_solicitacao": "2026-08-01",
            "numero": "05/2026",
            "descricao_evento": "Inauguração da Delegacia Cidadã",
            "quantidade": "40",
            "data_inicio_evento": "",
            "data_fim_evento": "",
            "periodo_evento_texto": "",
            "numero_nota_fiscal": "",
            "protocolo_pagamento": "",
            "data_atesto_gaf": "",
            "data_ordem_bancaria": "",
            "data_envio_empresa": "",
            "observacoes": "",
        }
        dados.update(kwargs)
        return dados

    def test_criacao_pela_view(self):
        resposta = self.client.post(
            reverse("coffee_break:nova"), self.dados_post()
        )
        solicitacao = SolicitacaoCoffeeBreak.objects.get(numero="05/2026")
        self.assertRedirects(
            resposta, reverse("coffee_break:detalhe", args=[solicitacao.pk])
        )
        self.assertEqual(solicitacao.criado_por, self.ascom)
        self.assertEqual(solicitacao.quantidade, 40)
        self.assertEqual(solicitacao.historico.count(), 1)

    def test_criacao_acima_do_saldo_mostra_erro(self):
        resposta = self.client.post(
            reverse("coffee_break:nova"), self.dados_post(quantidade="101")
        )
        self.assertEqual(resposta.status_code, 200)
        self.assertFalse(SolicitacaoCoffeeBreak.objects.exists())
        self.assertContains(resposta, "acima do saldo")

    def test_edicao_pela_view(self):
        solicitacao = self.criar_solicitacao(numero="05/2026")
        resposta = self.client.post(
            reverse("coffee_break:editar", args=[solicitacao.pk]),
            self.dados_post(
                quantidade="55",
                numero_nota_fiscal="8046",
                versao=str(int(solicitacao.atualizado_em.timestamp() * 1_000_000)),
            ),
        )
        solicitacao.refresh_from_db()
        self.assertRedirects(
            resposta, reverse("coffee_break:detalhe", args=[solicitacao.pk])
        )
        self.assertEqual(solicitacao.quantidade, 55)
        self.assertEqual(solicitacao.numero_nota_fiscal, "8046")
        self.assertEqual(solicitacao.historico.count(), 1)

    def test_edicao_concorrente_e_rejeitada(self):
        solicitacao = self.criar_solicitacao(numero="05/2026")
        versao_antiga = str(int(solicitacao.atualizado_em.timestamp() * 1_000_000))
        solicitacao.observacoes = "Alterada por outra pessoa"
        solicitacao.save(update_fields=["observacoes", "atualizado_em"])
        resposta = self.client.post(
            reverse("coffee_break:editar", args=[solicitacao.pk]),
            self.dados_post(versao=versao_antiga),
        )
        self.assertEqual(resposta.status_code, 200)
        self.assertContains(resposta, "alterada por outra pessoa")

    def test_concluida_nao_abre_edicao(self):
        solicitacao = self.criar_solicitacao(
            numero="05/2026",
            numero_nota_fiscal="1",
            protocolo_pagamento="P1",
            data_atesto_gaf=dt.date(2026, 8, 2),
            data_ordem_bancaria=dt.date(2026, 8, 3),
            data_envio_empresa=dt.date(2026, 8, 4),
        )
        resposta = self.client.get(
            reverse("coffee_break:editar", args=[solicitacao.pk])
        )
        self.assertRedirects(
            resposta, reverse("coffee_break:detalhe", args=[solicitacao.pk])
        )

    def test_form_nao_aceita_lote_inativo(self):
        lote_inativo = LoteCoffeeBreak.objects.create(
            contrato=self.contrato,
            numero=9,
            exercicio="2025",
            quantidade_total=500,
            ativo=False,
        )
        resposta = self.client.post(
            reverse("coffee_break:nova"), self.dados_post(lote=lote_inativo.pk)
        )
        self.assertEqual(resposta.status_code, 200)
        self.assertFalse(SolicitacaoCoffeeBreak.objects.exists())

    def test_busca_e_filtros_na_listagem(self):
        self.criar_solicitacao(descricao_evento="Curso de Power BI", numero="01/2026")
        self.criar_solicitacao(
            descricao_evento="Entrega de Medalhas",
            numero="02/2026",
            numero_nota_fiscal="8046",
        )
        url = reverse("coffee_break:solicitacoes")

        resposta = self.client.get(url, {"q": "Power"})
        self.assertContains(resposta, "Curso de Power BI")
        self.assertNotContains(resposta, "Entrega de Medalhas")

        resposta = self.client.get(url, {"q": "8046"})
        self.assertContains(resposta, "Entrega de Medalhas")

        resposta = self.client.get(
            url, {"situacao": SituacaoFinanceira.AGUARDANDO_PROTOCOLO}
        )
        self.assertContains(resposta, "Entrega de Medalhas")
        self.assertNotContains(resposta, "Curso de Power BI")

        resposta = self.client.get(url, {"lote": self.lote.pk, "q": "Power"})
        self.assertContains(resposta, "Curso de Power BI")

    def test_ordenacao_segura_ignora_campo_desconhecido(self):
        self.criar_solicitacao(numero="01/2026")
        resposta = self.client.get(
            reverse("coffee_break:solicitacoes"), {"ordem": "campo_malicioso"}
        )
        self.assertEqual(resposta.status_code, 200)

    def test_paginacao_preserva_filtros(self):
        for i in range(20):
            self.criar_solicitacao(quantidade=1, numero=f"{i + 1:02d}/2026")
        resposta = self.client.get(
            reverse("coffee_break:solicitacoes"), {"pagina": 2}
        )
        self.assertEqual(resposta.status_code, 200)
        self.assertContains(resposta, "pagina=1")

    def test_listagem_de_lotes_com_filtros(self):
        resposta = self.client.get(reverse("coffee_break:lotes"), {"q": "FAVO"})
        self.assertContains(resposta, "0762/2024")
        resposta = self.client.get(reverse("coffee_break:lotes"), {"q": "inexistente"})
        self.assertContains(resposta, "Nenhum lote encontrado")
        resposta = self.client.get(
            reverse("coffee_break:lotes"), {"situacao": "inativos"}
        )
        self.assertContains(resposta, "Nenhum lote encontrado")

    def test_detalhe_do_lote(self):
        self.criar_solicitacao(numero="01/2026")
        resposta = self.client.get(
            reverse("coffee_break:lote_detalhe", args=[self.lote.pk])
        )
        self.assertContains(resposta, "Lote 1 (2026)")
        self.assertContains(resposta, "0762/2024")
        self.assertContains(resposta, "01/2026")

    def test_painel_mostra_resumo(self):
        self.criar_solicitacao(quantidade=90)
        resposta = self.client.get(reverse("coffee_break:painel"))
        self.assertContains(resposta, "Capacidade contratada")
        self.assertContains(resposta, "restam")  # alerta de saldo baixo

    def test_cancelar_exige_post(self):
        solicitacao = self.criar_solicitacao()
        url = reverse("coffee_break:cancelar", args=[solicitacao.pk])
        self.assertEqual(self.client.get(url).status_code, 405)
        resposta = self.client.post(url, {"motivo": "Adiado"})
        self.assertRedirects(
            resposta, reverse("coffee_break:detalhe", args=[solicitacao.pk])
        )
        solicitacao.refresh_from_db()
        self.assertTrue(solicitacao.cancelada)

    def test_cancelar_exige_motivo(self):
        solicitacao = self.criar_solicitacao()
        resposta = self.client.post(
            reverse("coffee_break:cancelar", args=[solicitacao.pk]), {"motivo": ""}
        )
        self.assertRedirects(
            resposta, reverse("coffee_break:detalhe", args=[solicitacao.pk])
        )
        solicitacao.refresh_from_db()
        self.assertFalse(solicitacao.cancelada)

    def test_concluida_nao_pode_ser_cancelada(self):
        solicitacao = self.criar_solicitacao(
            numero_nota_fiscal="1",
            protocolo_pagamento="P1",
            data_atesto_gaf=dt.date(2026, 8, 2),
            data_ordem_bancaria=dt.date(2026, 8, 3),
            data_envio_empresa=dt.date(2026, 8, 4),
        )
        resposta = self.client.post(
            reverse("coffee_break:cancelar", args=[solicitacao.pk]),
            {"motivo": "Tentativa indevida"},
        )
        self.assertRedirects(
            resposta, reverse("coffee_break:detalhe", args=[solicitacao.pk])
        )
        solicitacao.refresh_from_db()
        self.assertFalse(solicitacao.cancelada)
        detalhe = self.client.get(reverse("coffee_break:detalhe", args=[solicitacao.pk]))
        self.assertNotContains(detalhe, "id_motivo_cancelamento")


class CadastrosCoffeeBreakTests(BaseCoffeeBreakTestCase):
    def test_operador_comum_nao_acessa_cadastros_contratuais(self):
        self.client.force_login(self.ascom)
        resposta = self.client.get(
            reverse("coffee_break:cadastro_lista", args=["fornecedores"])
        )
        self.assertEqual(resposta.status_code, 403)
        self.assertNotContains(
            self.client.get(reverse("coffee_break:painel")),
            reverse("coffee_break:cadastros"),
        )

    def test_administrador_cria_fornecedor_pela_interface(self):
        self.client.force_login(self.admin_modulo)
        resposta = self.client.post(
            reverse("coffee_break:cadastro_novo", args=["fornecedores"]),
            {
                "razao_social": "Fornecedor Institucional Ltda",
                "cnpj": "12.345.678/0001-95",
                "contato": "Equipe comercial",
                "telefone": "41 3333-4444",
                "email": "contato@example.com",
                "ativo": "1",
            },
        )
        self.assertRedirects(
            resposta,
            reverse("coffee_break:cadastro_lista", args=["fornecedores"]),
        )
        self.assertTrue(
            Fornecedor.objects.filter(
                razao_social="Fornecedor Institucional Ltda",
                cnpj="12345678000195",
            ).exists()
        )
        painel = self.client.get(reverse("coffee_break:painel"))
        self.assertContains(painel, reverse("coffee_break:cadastros"))

    def test_capacidade_do_lote_nao_pode_ficar_abaixo_do_consumido(self):
        self.criar_solicitacao(quantidade=30)
        self.client.force_login(self.admin_modulo)
        resposta = self.client.post(
            reverse("coffee_break:cadastro_editar", args=["lotes", self.lote.pk]),
            {
                "contrato": self.contrato.pk,
                "numero": self.lote.numero,
                "exercicio": self.lote.exercicio,
                "quantidade_total": 20,
                "empenho": self.lote.empenho,
                "municipios_texto": "",
                "orientacoes": "",
                "especificacoes_tecnicas": "",
                "observacoes": "",
                "ativo": "1",
                "versao": str(int(self.lote.atualizado_em.timestamp() * 1_000_000)),
            },
        )
        self.assertEqual(resposta.status_code, 200)
        self.assertContains(resposta, "já consumiu 30 unidades")


# ---------------------------------------------------------------------------
# Importação da planilha
# ---------------------------------------------------------------------------

def _planilha_de_teste():
    """Monta em memória uma planilha mínima no formato da original."""
    import openpyxl

    wb = openpyxl.Workbook()
    geral = wb.active
    geral.title = "Geral"
    geral["F2"] = "Orientações gerais de teste"
    geral["J3"] = "E" * 150  # especificações técnicas (texto longo)

    aba = wb.create_sheet(" LOTE 1- 2026")
    aba["A1"] = "COFFEE BREAK- ASCOM - LOTE 1 - FAVO E MEL (CONTRATO 0762/2024)"
    aba["A2"] = "CNPJ: 35014719000166 - EMPENHO: 2026NE030208"
    aba["J1"] = "TOTAL"
    aba["J3"] = 3800
    aba["L3"] = "CIDADES ABRANGENTES: Cidade Modelo, Desconhecida, Vila Exemplo e Cidade Par."
    aba["A3"] = "Data da Solicitação"
    aba["B3"] = "Data do evento"
    aba["C3"] = "N° da Solicitação "
    aba["D3"] = "Descrição do evento"
    aba["E3"] = "Número da Nota Fiscal"
    aba["F3"] = "Quantidade "
    aba["G3"] = "N° protocolo de pagamento"
    aba["H3"] = "Data de atesto e envio ao GAF "
    # Linha normal: nº de solicitação convertido em data pelo Excel.
    aba["A4"] = dt.datetime(2026, 2, 10)
    aba["B4"] = dt.datetime(2026, 2, 11)
    aba["C4"] = dt.datetime(2026, 2, 1)  # "02/2026"
    aba["D4"] = "Central de Flagrantes"
    aba["E4"] = 8046
    aba["F4"] = 40
    aba["G4"] = " 25.419.856-0 "
    aba["H4"] = dt.datetime(2026, 2, 18)
    # Linha com período textual e sem NF.
    aba["A5"] = dt.datetime(2026, 3, 18)
    aba["B5"] = "23,  24 e 25/03"
    aba["C5"] = "13/2026"
    aba["D5"] = "Encontro Nacional"
    aba["F5"] = 500
    # Linha cancelada.
    aba["A6"] = dt.datetime(2026, 6, 22)
    aba["B6"] = dt.datetime(2026, 6, 30)
    aba["C6"] = "16/2026"
    aba["D6"] = 'Projeto "Rota de Proteção" CANCELADO'
    aba["E6"] = "-"
    aba["F6"] = 0
    aba["G6"] = "-"
    aba["H6"] = "-"

    ob = wb.create_sheet("Controle de Ordem Bancária")
    ob["A1"] = "LOTE 1 - FAVO E MEL -  CNPJ: 35014719000166 - EMPENHO: 2026NE030208"
    ob["A2"] = "N° da Solicitação "
    ob["B2"] = "Descrição do evento"
    ob["C2"] = "Número da NF"
    ob["D2"] = "N° protocolo de pagamento"
    ob["E2"] = "OB emitida em"
    ob["F2"] = "Data de envio p/ empresa"
    ob["A3"] = dt.datetime(2026, 2, 1)  # "02/2026"
    ob["B3"] = "Central de Flagrantes"
    ob["C3"] = 8046
    ob["D3"] = " 25.419.856-0 "
    ob["E3"] = dt.datetime(2026, 2, 24)
    ob["F3"] = dt.datetime(2026, 6, 16)
    return wb


class ImportacaoTests(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.usuario = User.objects.create_superuser("importador", password="x")
        regiao = Regiao.objects.create(nome="Região Teste")
        estado = Estado.objects.get(codigo_ibge=41)
        Municipio.objects.create(nome="Cidade Modelo", estado=estado, regiao=regiao)
        Municipio.objects.create(nome="Vila Exemplo", estado=estado, regiao=regiao)
        Municipio.objects.create(nome="Cidade Par", estado=estado, regiao=regiao)

    def setUp(self):
        import tempfile

        wb = _planilha_de_teste()
        self.arquivo = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(self.arquivo.name)
        self.arquivo.close()

    def tearDown(self):
        import os

        os.unlink(self.arquivo.name)

    def importar(self, *flags):
        saida = io.StringIO()
        call_command(
            "importar_coffee_break",
            self.arquivo.name,
            "--usuario",
            "importador",
            *flags,
            stdout=saida,
        )
        return saida.getvalue()

    def test_importacao_cria_estruturas_e_solicitacoes(self):
        self.importar()
        lote = LoteCoffeeBreak.objects.get(numero=1, exercicio="2026")
        self.assertEqual(lote.quantidade_total, 3800)
        self.assertEqual(lote.empenho, "2026NE030208")
        self.assertEqual(lote.contrato.numero, "0762/2024")
        self.assertEqual(
            lote.contrato.fornecedor.razao_social,
            "PADARIA E CONFEITARIA FAVO E MEL LTDA",
        )
        # Municípios reconhecidos, inclusive o par final "Vila Exemplo e
        # Cidade Par"; "Desconhecida" fica apenas no texto original.
        nomes = set(lote.municipios.values_list("nome", flat=True))
        self.assertEqual(nomes, {"Cidade Modelo", "Vila Exemplo", "Cidade Par"})
        self.assertIn("CIDADES ABRANGENTES", lote.municipios_texto)
        self.assertIn("Desconhecida", lote.municipios_texto)

        normal = SolicitacaoCoffeeBreak.objects.get(lote=lote, numero="02/2026")
        self.assertEqual(normal.quantidade, 40)
        self.assertEqual(normal.numero_nota_fiscal, "8046")
        self.assertEqual(normal.protocolo_pagamento, "25.419.856-0")
        self.assertEqual(normal.data_atesto_gaf, dt.date(2026, 2, 18))
        self.assertFalse(normal.cancelada)

        textual = SolicitacaoCoffeeBreak.objects.get(lote=lote, numero="13/2026")
        self.assertEqual(textual.periodo_evento_texto, "23, 24 e 25/03")
        self.assertEqual(textual.data_inicio_evento, dt.date(2026, 3, 23))
        self.assertEqual(textual.data_fim_evento, dt.date(2026, 3, 25))

        cancelada = SolicitacaoCoffeeBreak.objects.get(lote=lote, numero="16/2026")
        self.assertTrue(cancelada.cancelada)
        self.assertEqual(cancelada.quantidade, 0)
        # Cancelada fora do consumo.
        self.assertEqual(lote.quantidade_consumida, 540)

    def test_importa_marcos_da_ordem_bancaria(self):
        self.importar()
        solicitacao = SolicitacaoCoffeeBreak.objects.get(numero="02/2026")
        self.assertEqual(solicitacao.data_ordem_bancaria, dt.date(2026, 2, 24))
        self.assertEqual(solicitacao.data_envio_empresa, dt.date(2026, 6, 16))
        self.assertEqual(
            solicitacao.situacao_financeira, SituacaoFinanceira.CONCLUIDA
        )

    def test_importacao_idempotente(self):
        self.importar()
        contagens = (
            Fornecedor.objects.count(),
            ContratoCoffeeBreak.objects.count(),
            LoteCoffeeBreak.objects.count(),
            SolicitacaoCoffeeBreak.objects.count(),
        )
        self.importar()
        self.assertEqual(
            contagens,
            (
                Fornecedor.objects.count(),
                ContratoCoffeeBreak.objects.count(),
                LoteCoffeeBreak.objects.count(),
                SolicitacaoCoffeeBreak.objects.count(),
            ),
        )

    def test_dry_run_nao_persiste(self):
        saida = self.importar("--dry-run")
        self.assertIn("[dry-run]", saida)
        self.assertFalse(Fornecedor.objects.exists())
        self.assertFalse(SolicitacaoCoffeeBreak.objects.exists())

    def test_identificador_como_texto(self):
        self.assertEqual(
            identificador_como_texto(dt.datetime(2026, 2, 1)), "02/2026"
        )
        self.assertEqual(identificador_como_texto(8046.0), "8046")
        self.assertEqual(identificador_como_texto(" 13/2026 "), "13/2026")
        self.assertEqual(identificador_como_texto("-"), "")
        self.assertEqual(identificador_como_texto(None), "")

    def test_parse_periodo_livre(self):
        self.assertEqual(
            parse_periodo_livre("23,  24 e 25/03", 2026),
            (dt.date(2026, 3, 23), dt.date(2026, 3, 25)),
        )
        self.assertEqual(
            parse_periodo_livre("12 à 15/05", 2026),
            (dt.date(2026, 5, 12), dt.date(2026, 5, 15)),
        )
        self.assertEqual(
            parse_periodo_livre("03 e 10/11/2025", 2026),
            (dt.date(2025, 11, 3), dt.date(2025, 11, 10)),
        )
        # "17 e 20/2026": mês 20 não existe — fica só como texto.
        self.assertEqual(parse_periodo_livre("17 e 20/2026", 2026), (None, None))


# ---------------------------------------------------------------------------
# Administração
# ---------------------------------------------------------------------------

class AdminTests(BaseCoffeeBreakTestCase):
    def setUp(self):
        self.client.force_login(self.superusuario)

    def test_listas_do_admin_respondem(self):
        for rota in (
            "admin:coffee_break_fornecedor_changelist",
            "admin:coffee_break_contratocoffeebreak_changelist",
            "admin:coffee_break_lotecoffeebreak_changelist",
            "admin:coffee_break_solicitacaocoffeebreak_changelist",
            "admin:accounts_setor_changelist",
            "admin:accounts_modulo_changelist",
        ):
            self.assertEqual(self.client.get(reverse(rota)).status_code, 200, rota)

    def test_admin_bloqueia_capacidade_abaixo_do_consumido(self):
        self.criar_solicitacao(quantidade=80)
        resposta = self.client.post(
            reverse("admin:coffee_break_lotecoffeebreak_change", args=[self.lote.pk]),
            {
                "contrato": self.contrato.pk,
                "numero": 1,
                "exercicio": "2026",
                "quantidade_total": 50,  # abaixo dos 80 consumidos
                "empenho": "2026NE030208",
                "municipios_texto": "",
                "orientacoes": "",
                "especificacoes_tecnicas": "",
                "observacoes": "",
                "ativo": "on",
            },
        )
        self.assertEqual(resposta.status_code, 200)
        self.assertContains(resposta, "não pode ficar abaixo")
        self.lote.refresh_from_db()
        self.assertEqual(self.lote.quantidade_total, 100)

    def test_admin_bloqueia_solicitacao_acima_do_saldo(self):
        resposta = self.client.post(
            reverse("admin:coffee_break_solicitacaocoffeebreak_add"),
            {
                "lote": self.lote.pk,
                "data_solicitacao": "2026-08-01",
                "numero": "01/2026",
                "descricao_evento": "Estouro",
                "quantidade": 101,
                "numero_nota_fiscal": "",
                "protocolo_pagamento": "",
                "observacoes": "",
                "motivo_cancelamento": "",
                "criado_por": self.superusuario.pk,
            },
        )
        self.assertEqual(resposta.status_code, 200)
        self.assertContains(resposta, "acima do saldo")
        self.assertFalse(SolicitacaoCoffeeBreak.objects.exists())
