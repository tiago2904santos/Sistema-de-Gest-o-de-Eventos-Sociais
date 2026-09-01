import tempfile
from datetime import date, datetime
from pathlib import Path

from django.contrib.auth import get_user_model
from django.core.management import call_command
from django.test import TestCase
from django.urls import reverse
from openpyxl import Workbook

from accounts.models import Modulo, Setor
from cadastros.models import TipoEvento

from .forms import DemandaEventoForm
from .models import (
    DemandaEvento,
    Palestrante,
    RespostaPadrao,
    StatusDemanda,
    Tema,
)
from .permissions import CODIGO_MODULO

User = get_user_model()


class BaseDemandasTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.modulo = Modulo.objects.get(codigo=CODIGO_MODULO)
        cls.ascom = Setor.objects.get(sigla="ASCOM")
        cls.usuario = User.objects.create_user("ascom", password="x")
        cls.usuario.setores.add(cls.ascom)
        cls.outro_setor = Setor.objects.create(sigla="ASCOM-2", nome="Outra equipe ASCOM")
        cls.outro_setor.modulos.add(cls.modulo)
        cls.outro = User.objects.create_user("outra_ascom", password="x")
        cls.outro.setores.add(cls.outro_setor)
        cls.sem_modulo = User.objects.create_user("iipr", password="x")
        cls.sem_modulo.setores.add(Setor.objects.create(sigla="IIPR", nome="Instituto de Identificação"))
        cls.superusuario = User.objects.create_superuser("root_demandas", password="x")
        cls.tipo = TipoEvento.objects.get_or_create(nome="Palestra")[0]
        cls.tema = Tema.objects.create(nome="Crimes virtuais")

    def criar_demanda(self, setor=None, **kwargs):
        dados = {
            "data_solicitacao": date(2026, 8, 1),
            "tipo_evento": self.tipo,
            "tema": self.tema,
            "solicitante": "Escola Municipal",
            "status": StatusDemanda.PENDENTE,
            "criado_por": self.usuario,
        }
        dados.update(kwargs)
        demanda = DemandaEvento.objects.create(**dados)
        demanda.setores.add(setor or self.ascom)
        return demanda


class AcessoDemandasTests(BaseDemandasTestCase):
    def test_modulo_bloqueia_url_direta(self):
        self.client.force_login(self.sem_modulo)
        self.assertEqual(self.client.get(reverse("demandas_eventos:dashboard")).status_code, 403)

    def test_portal_mostra_modulo_conforme_acesso(self):
        self.client.force_login(self.usuario)
        resposta = self.client.get(reverse("core:home"))
        self.assertContains(resposta, reverse("demandas_eventos:dashboard"))
        self.client.force_login(self.sem_modulo)
        resposta = self.client.get(reverse("core:home"))
        self.assertNotContains(resposta, reverse("demandas_eventos:dashboard"))

    def test_navbar_contextual_dentro_do_modulo(self):
        self.client.force_login(self.usuario)
        resposta = self.client.get(reverse("demandas_eventos:dashboard"))
        self.assertContains(resposta, f'href="{reverse("demandas_eventos:lista")}"')
        self.assertNotContains(resposta, f'href="{reverse("coffee_break:painel")}"')

    def test_isolamento_por_setor_e_superusuario_global(self):
        demanda = self.criar_demanda()
        self.client.force_login(self.outro)
        self.assertNotContains(self.client.get(reverse("demandas_eventos:lista")), "Escola Municipal")
        self.assertEqual(self.client.get(reverse("demandas_eventos:detalhe", args=[demanda.pk])).status_code, 404)
        self.client.force_login(self.superusuario)
        self.assertEqual(self.client.get(reverse("demandas_eventos:detalhe", args=[demanda.pk])).status_code, 200)


class FormulariosViewsTests(BaseDemandasTestCase):
    def dados_post(self):
        return {
            "data_solicitacao": "2026-08-20",
            "tipo_evento": self.tipo.pk,
            "tema": self.tema.pk,
            "canal_solicitacao": "E-mail",
            "solicitante": "Colégio Estadual",
            "contato": "colegio@example.org",
            "data_inicio_evento": "2026-09-10",
            "data_fim_evento": "2026-09-11",
            "periodo_evento_texto": "",
            "setores": [self.ascom.pk],
        }

    def test_form_rejeita_periodo_invertido(self):
        dados = self.dados_post()
        dados["data_fim_evento"] = "2026-09-01"
        form = DemandaEventoForm(dados, usuario=self.usuario)
        self.assertFalse(form.is_valid())
        self.assertIn("data_fim_evento", form.errors)

    def test_solicitante_com_texto_legado_extenso_e_preservado(self):
        self.assertEqual(DemandaEvento._meta.get_field("solicitante").max_length, 1000)
        demanda = self.criar_demanda(solicitante="A" * 500)
        self.assertEqual(len(demanda.solicitante), 500)

    def test_cria_edita_e_exibe_demanda_reutilizando_componentes(self):
        self.client.force_login(self.usuario)
        resposta = self.client.post(reverse("demandas_eventos:nova"), self.dados_post())
        demanda = DemandaEvento.objects.get(solicitante="Colégio Estadual")
        self.assertRedirects(resposta, reverse("demandas_eventos:detalhe", args=[demanda.pk]))
        resposta = self.client.get(reverse("demandas_eventos:detalhe", args=[demanda.pk]))
        self.assertContains(resposta, "Colégio Estadual")
        self.assertContains(resposta, "Demanda #")
        self.assertContains(resposta, "Pendente")
        self.assertEqual(demanda.historico.count(), 1)

        resposta = self.client.post(
            reverse("demandas_eventos:transicionar", args=[demanda.pk]),
            {"novo_status": StatusDemanda.EM_ANDAMENTO},
        )
        self.assertRedirects(
            resposta, reverse("demandas_eventos:detalhe", args=[demanda.pk])
        )
        demanda.refresh_from_db()
        self.assertEqual(demanda.status, StatusDemanda.EM_ANDAMENTO)
        self.assertEqual(demanda.historico.count(), 2)

    def test_responsavel_precisa_pertencer_ao_setor_envolvido(self):
        dados = self.dados_post()
        dados["responsavel_atendimento"] = self.outro.pk
        form = DemandaEventoForm(dados, usuario=self.usuario)
        self.assertFalse(form.is_valid())
        self.assertIn("responsavel_atendimento", form.errors)

    def test_status_final_exige_justificativa_e_bloqueia_edicao(self):
        demanda = self.criar_demanda(status=StatusDemanda.EM_ANDAMENTO)
        self.client.force_login(self.usuario)
        resposta = self.client.post(
            reverse("demandas_eventos:transicionar", args=[demanda.pk]),
            {"novo_status": StatusDemanda.CANCELADA, "justificativa": ""},
        )
        demanda.refresh_from_db()
        self.assertEqual(demanda.status, StatusDemanda.EM_ANDAMENTO)

        self.client.post(
            reverse("demandas_eventos:transicionar", args=[demanda.pk]),
            {"novo_status": StatusDemanda.CANCELADA, "justificativa": "Evento cancelado."},
        )
        demanda.refresh_from_db()
        self.assertEqual(demanda.status, StatusDemanda.CANCELADA)
        self.assertEqual(
            self.client.get(reverse("demandas_eventos:editar", args=[demanda.pk])).status_code,
            403,
        )

    def test_edicao_concorrente_e_rejeitada(self):
        demanda = self.criar_demanda()
        versao_antiga = str(int(demanda.atualizado_em.timestamp() * 1_000_000))
        demanda.andamento = "Alteração concorrente"
        demanda.save(update_fields=["andamento", "atualizado_em"])
        dados = self.dados_post()
        dados["versao"] = versao_antiga
        self.client.force_login(self.usuario)
        resposta = self.client.post(
            reverse("demandas_eventos:editar", args=[demanda.pk]), dados
        )
        self.assertEqual(resposta.status_code, 200)
        self.assertContains(resposta, "alterada por outra pessoa")

    def test_listagem_filtra_status_e_busca(self):
        self.criar_demanda(solicitante="Demanda visível", status=StatusDemanda.PENDENTE)
        self.criar_demanda(solicitante="Demanda concluída", status=StatusDemanda.ATENDIDA)
        self.client.force_login(self.usuario)
        resposta = self.client.get(reverse("demandas_eventos:lista"), {"status": StatusDemanda.ATENDIDA})
        self.assertContains(resposta, "Demanda concluída")
        self.assertNotContains(resposta, "Demanda visível")
        resposta = self.client.get(reverse("demandas_eventos:lista"), {"q": "visível"})
        self.assertContains(resposta, "Demanda visível")

    def test_datas_invalidas_nos_filtros_nao_geram_erro(self):
        self.criar_demanda(solicitante="Demanda visível")
        self.client.force_login(self.usuario)
        resposta = self.client.get(
            reverse("demandas_eventos:lista"),
            {"inicio": "31-02-2026", "fim": "invalida"},
        )
        self.assertEqual(resposta.status_code, 200)
        resposta = self.client.get(
            reverse("demandas_eventos:exportar"), {"inicio": "invalida"}
        )
        self.assertEqual(resposta.status_code, 200)

    def test_exportacao_respeita_recorte_visivel(self):
        self.criar_demanda(solicitante="Demanda exportável")
        self.criar_demanda(setor=self.outro_setor, solicitante="Demanda de outro setor")
        self.client.force_login(self.usuario)
        resposta = self.client.get(reverse("demandas_eventos:exportar"))
        conteudo = resposta.content.decode("utf-8-sig")
        self.assertIn("Demanda exportável", conteudo)
        self.assertNotIn("Demanda de outro setor", conteudo)

    def test_cadastro_de_resposta_padrao(self):
        self.client.force_login(self.usuario)
        resposta = self.client.post(
            reverse("demandas_eventos:cadastro_novo", args=["respostas"]),
            {"tipo": "Pedido incompleto", "mensagem": "Solicite os dados faltantes.", "ativo": "1"},
        )
        self.assertEqual(resposta.status_code, 302)
        self.assertTrue(RespostaPadrao.objects.filter(tipo="Pedido incompleto").exists())


class ImportacaoPlanilhaTests(BaseDemandasTestCase):
    def setUp(self):
        super().setUp()
        self.tmp = tempfile.TemporaryDirectory()
        self.arquivo = Path(self.tmp.name) / "ascom.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "2026"
        ws.append([
            "MÊS", "MUNICIPIO", "DATA DO EVENTO E HORA (PERÍODO)", "EVENTO",
            "STATUS DA DEMANDA", "ANDAMENTO", "INFORMAÇÕES PRÉVIAS", "SOLICITANTE",
            "CONTATO", "DATA DA SOLICITAÇÃO", "FOI SOLICITADO VIA:", "DESCRIÇÃO",
            "QUANTIDADE DE PÚBLICO", "ASSUNTO E-MAIL", "PEDIDO/CONTATO",
        ])
        ws.append([
            "AGOSTO", "Curitiba", "10 e 11/09", "Palestra", "EVENTO AGENDADO",
            "Confirmado", "Levar projetor", "Escola Teste", "41 99999-0000",
            datetime(2026, 8, 20), "E-MAIL", "Palestra educativa", 120,
            "Solicitação de palestra", "Pedido completo",
        ])
        temas = wb.create_sheet("TEMAS")
        temas.append(["Prevenção"])
        palestrantes = wb.create_sheet("PALESTRANTES")
        palestrantes.append(["MUNICÍPIO", "DIVISÃO", "LOTAÇÃO", "SERVIDOR", "CONTATO", "E-MAIL", "TEMA DE ABORDAGEM"])
        palestrantes.append(["Curitiba", "DIC", "NUCIBER", "Dra. Teste", "41 90000-0000", "teste@pc.pr.gov.br", "Prevenção"])
        respostas = wb.create_sheet("Repostas Padrão")
        respostas.append(["TIPO", "MENSAGEM"])
        respostas.append(["Recebimento", "Demanda recebida."])
        wb.save(self.arquivo)
        wb.close()

    def tearDown(self):
        self.tmp.cleanup()
        super().tearDown()

    def test_dry_run_nao_persiste(self):
        call_command("importar_planilha_ascom", str(self.arquivo), dry_run=True)
        self.assertFalse(DemandaEvento.objects.filter(solicitante="Escola Teste").exists())

    def test_importacao_idempotente_e_preserva_periodo_textual(self):
        call_command("importar_planilha_ascom", str(self.arquivo))
        call_command("importar_planilha_ascom", str(self.arquivo))
        self.assertEqual(DemandaEvento.objects.filter(solicitante="Escola Teste").count(), 1)
        demanda = DemandaEvento.objects.get(solicitante="Escola Teste")
        self.assertEqual(demanda.periodo_evento_texto, "10 e 11/09")
        self.assertEqual(demanda.status, StatusDemanda.EVENTO_AGENDADO)
        self.assertEqual(demanda.historico.count(), 1)
        self.assertTrue(Palestrante.objects.filter(nome="Dra. Teste").exists())
        self.assertTrue(RespostaPadrao.objects.filter(tipo="Recebimento").exists())
