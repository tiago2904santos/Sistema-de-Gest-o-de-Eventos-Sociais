"""Importa o histórico de Palestras e Eventos da ASCOM sem duplicar registros."""

import hashlib
import re
import unicodedata
from datetime import date, datetime
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

from accounts.models import Setor
from cadastros.models import Municipio, TipoEvento
from demandas_eventos.models import (
    DemandaEvento,
    Palestrante,
    RespostaPadrao,
    StatusDemanda,
    Tema,
)


def _texto(valor):
    if valor is None:
        return ""
    return str(valor).strip()


def _chave(valor):
    valor = unicodedata.normalize("NFKD", _texto(valor)).encode("ascii", "ignore").decode()
    return re.sub(r"[^A-Z0-9]+", " ", valor.upper()).strip()


def _data(valor):
    if isinstance(valor, datetime):
        return valor.date()
    if isinstance(valor, date):
        return valor
    texto = _texto(valor)
    for formato in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(texto, formato).date()
        except ValueError:
            continue
    return None


STATUS = {
    "PENDENTE": StatusDemanda.PENDENTE,
    "AGUARDANDO RETORNO": StatusDemanda.AGUARDANDO_RETORNO,
    "EM ANDAMENTO": StatusDemanda.EM_ANDAMENTO,
    "EVENTO AGENDADO": StatusDemanda.EVENTO_AGENDADO,
    "PALESTRA AGENDADA": StatusDemanda.EVENTO_AGENDADO,
    "ATENDIDA": StatusDemanda.ATENDIDA,
    "SOL ATENDIDA": StatusDemanda.ATENDIDA,
    "NAO ATENDER": StatusDemanda.NAO_ATENDER,
    "CANCELADA": StatusDemanda.CANCELADA,
}


MAPAS = {
    "2022": {
        "data_solicitacao": "Data", "solicitante": "Solicitante", "contato": "Contato",
        "tipo": "Evento", "tema": "TEMA", "pedido": "Pedido/Contato",
        "periodo": "Data do evento", "status": "Status da demanda",
        "responsavel": "Responsável pelo atendimento", "andamento": "Andamento",
        "servidor": "Servidor", "unidade": "Unidade", "publico": "Público",
        "briefing": "Briefing", "materia": "Texto publicado na página",
    },
    "2023": {
        "data_solicitacao": "Data", "solicitante": "Solicitante", "contato": "Contato",
        "tipo": "Evento", "organizacao": "Responsável pela organização", "tema": "TEMA",
        "pedido": "Pedido/Contato", "periodo": "Data do evento", "status": "Status da demanda",
        "responsavel": "Responsável pelo atendimento", "andamento": "Andamento",
        "servidor": "Servidor", "unidade": "Unidade", "publico": "Público",
        "briefing": "Briefing", "materia": "Texto publicado na página",
    },
    "2024": {
        "data_solicitacao": "Data da Solicitação", "tipo": "Tipo de Evento",
        "canal": "Por onde foi solicitado", "tema": "Tema", "status": "Status da demanda",
        "andamento": "Andamento", "periodo": "Data do evento", "solicitante": "Solicitante",
        "contato": "Contato", "municipio": "Município", "pedido": "Pedido/Contato",
        "servidor": "Servidor", "unidade": "Unidade", "publico": "Público",
        "briefing": "Breafing Palestra", "materia": "Matéria no Site",
        "responsavel": "Responsável pelo Atendimento",
    },
    "2025": {
        "data_solicitacao": "Data da Solicitação", "tipo": "EVENTO",
        "canal": "Por onde foi solicitado", "tema": "Tema", "publico": "Quantidade de público",
        "status": "Status da demanda", "andamento": "Andamento", "municipio": "Município",
        "periodo": "Data do evento", "solicitante": "Solicitante", "contato": "Contato",
        "assunto": "ASSUNTO E-MAIL", "pedido": "Pedido/Contato", "servidor": "Servidor",
    },
    "2026": {
        "municipio": "MUNICIPIO", "periodo": "DATA DO EVENTO E HORA (PERÍODO)",
        "tipo": "EVENTO", "status": "STATUS DA DEMANDA", "andamento": "ANDAMENTO",
        "informacoes": "INFORMAÇÕES PRÉVIAS", "solicitante": "SOLICITANTE",
        "contato": "CONTATO", "data_solicitacao": "DATA DA SOLICITAÇÃO",
        "canal": "FOI SOLICITADO VIA:", "descricao": "DESCRIÇÃO",
        "publico": "QUANTIDADE DE PÚBLICO", "assunto": "ASSUNTO E-MAIL",
        "pedido": "PEDIDO/CONTATO",
    },
}


class Command(BaseCommand):
    help = "Importa a planilha Palestras e Eventos ASCOM (.xlsx)"

    def add_arguments(self, parser):
        parser.add_argument("arquivo")
        parser.add_argument("--dry-run", action="store_true", help="Valida sem persistir")

    @transaction.atomic
    def handle(self, *args, **options):
        caminho = Path(options["arquivo"])
        if not caminho.is_file():
            raise CommandError(f"Arquivo não encontrado: {caminho}")
        ascom = Setor.objects.filter(sigla="ASCOM").first()
        if not ascom:
            raise CommandError("O setor ASCOM não está cadastrado. Aplique as migrations da fundação.")

        workbook = load_workbook(caminho, data_only=True, read_only=False)
        resumo = {"demandas": 0, "palestrantes": 0, "temas": 0, "respostas": 0, "ignoradas": 0, "avisos": 0}
        try:
            self._importar_temas(workbook, resumo)
            self._importar_palestrantes(workbook, resumo)
            self._importar_respostas(workbook, resumo)
            for aba, mapa in MAPAS.items():
                if aba in workbook.sheetnames:
                    self._importar_demandas(workbook[aba], aba, mapa, ascom, resumo)
        finally:
            workbook.close()

        if options["dry_run"]:
            transaction.set_rollback(True)
            prefixo = "Dry-run concluído"
        else:
            prefixo = "Importação concluída"
        self.stdout.write(self.style.SUCCESS(prefixo + ": " + ", ".join(f"{k}={v}" for k, v in resumo.items())))

    def _importar_temas(self, workbook, resumo):
        if "TEMAS" not in workbook.sheetnames:
            return
        for (valor,) in workbook["TEMAS"].iter_rows(values_only=True, max_col=1):
            nome = _texto(valor)
            if not nome:
                continue
            _, criado = Tema.objects.get_or_create(nome__iexact=nome, defaults={"nome": nome})
            resumo["temas"] += int(criado)

    def _importar_palestrantes(self, workbook, resumo):
        if "PALESTRANTES" not in workbook.sheetnames:
            return
        ws = workbook["PALESTRANTES"]
        headers = {_chave(c.value): i for i, c in enumerate(ws[1]) if c.value}
        for valores in ws.iter_rows(min_row=2, values_only=True):
            nome = _texto(valores[headers.get("SERVIDOR", -1)]) if "SERVIDOR" in headers else ""
            if not nome:
                continue
            lotacao = _texto(valores[headers.get("LOTACAO", -1)]) if "LOTACAO" in headers else ""
            defaults = {
                "municipio_texto": _texto(valores[headers.get("MUNICIPIO", -1)]) if "MUNICIPIO" in headers else "",
                "divisao": _texto(valores[headers.get("DIVISAO", -1)]) if "DIVISAO" in headers else "",
                "contato": _texto(valores[headers.get("CONTATO", -1)]) if "CONTATO" in headers else "",
                "email": _texto(valores[headers.get("E MAIL", -1)]) if "E MAIL" in headers else "",
            }
            palestrante, criado = Palestrante.objects.update_or_create(nome=nome, lotacao=lotacao, defaults=defaults)
            resumo["palestrantes"] += int(criado)
            tema_nome = _texto(valores[headers.get("TEMA DE ABORDAGEM", -1)]) if "TEMA DE ABORDAGEM" in headers else ""
            if tema_nome:
                tema, tema_criado = Tema.objects.get_or_create(nome__iexact=tema_nome, defaults={"nome": tema_nome})
                palestrante.temas.add(tema)
                resumo["temas"] += int(tema_criado)

    def _importar_respostas(self, workbook, resumo):
        nome_aba = next((n for n in workbook.sheetnames if _chave(n) == "REPOSTAS PADRAO"), None)
        if not nome_aba:
            return
        for tipo, mensagem, *_ in workbook[nome_aba].iter_rows(min_row=2, values_only=True):
            tipo, mensagem = _texto(tipo), _texto(mensagem)
            if not tipo or not mensagem:
                continue
            _, criado = RespostaPadrao.objects.update_or_create(tipo=tipo, defaults={"mensagem": mensagem})
            resumo["respostas"] += int(criado)

    def _importar_demandas(self, ws, aba, mapa, ascom, resumo):
        headers = {_chave(c.value): i for i, c in enumerate(ws[1]) if c.value}

        def obter(valores, campo):
            cabecalho = mapa.get(campo)
            indice = headers.get(_chave(cabecalho)) if cabecalho else None
            return valores[indice] if indice is not None and indice < len(valores) else None

        for numero, valores in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            solicitante = _texto(obter(valores, "solicitante"))
            tipo_nome = _texto(obter(valores, "tipo"))
            pedido = _texto(obter(valores, "pedido"))
            if not any((solicitante, tipo_nome, pedido)):
                continue
            data_solicitacao = _data(obter(valores, "data_solicitacao"))
            if not data_solicitacao:
                resumo["avisos"] += 1
                self.stderr.write(
                    self.style.WARNING(
                        f"{aba}, linha {numero}: data da solicitação inválida; linha ignorada."
                    )
                )
                continue
            tipo_nome = tipo_nome or "Evento"
            tipo = TipoEvento.objects.filter(nome__iexact=tipo_nome).first()
            if not tipo:
                tipo = TipoEvento.objects.create(nome=tipo_nome)
            tema_nome = _texto(obter(valores, "tema"))
            tema = None
            if tema_nome:
                tema = Tema.objects.filter(nome__iexact=tema_nome).first()
                if not tema:
                    tema = Tema.objects.create(nome=tema_nome)
                    resumo["temas"] += 1
            municipio_nome = _texto(obter(valores, "municipio"))
            municipio = Municipio.objects.filter(nome__iexact=municipio_nome).first() if municipio_nome else None
            periodo_original = obter(valores, "periodo")
            data_evento = _data(periodo_original)
            periodo_texto = "" if data_evento else _texto(periodo_original)
            publico_texto = _texto(obter(valores, "publico"))
            try:
                publico = int(float(publico_texto)) if publico_texto and float(publico_texto) >= 0 else None
            except ValueError:
                publico = None
            status_original = _chave(obter(valores, "status"))
            status = STATUS.get(status_original, StatusDemanda.PENDENTE)
            if status_original and status_original not in STATUS:
                resumo["avisos"] += 1
                self.stderr.write(
                    self.style.WARNING(
                        f"{aba}, linha {numero}: status '{status_original}' não reconhecido; "
                        "importado como Pendente."
                    )
                )
            identidade = "|".join([aba, str(numero), str(data_solicitacao), solicitante, tipo_nome, pedido[:100]])
            chave = hashlib.sha256(identidade.encode("utf-8")).hexdigest()
            defaults = {
                "data_solicitacao": data_solicitacao,
                "tipo_evento": tipo,
                "tema": tema,
                "canal_solicitacao": _texto(obter(valores, "canal")),
                "municipio": municipio,
                "municipio_texto": "" if municipio else municipio_nome,
                "data_inicio_evento": data_evento,
                "periodo_evento_texto": periodo_texto,
                "solicitante": solicitante or "Não informado",
                "contato": _texto(obter(valores, "contato")),
                "assunto_email": _texto(obter(valores, "assunto")),
                "pedido_contato": pedido,
                "descricao": _texto(obter(valores, "descricao")),
                "status": status,
                "andamento": _texto(obter(valores, "andamento")),
                "informacoes_previas": _texto(obter(valores, "informacoes")),
                "responsavel_organizacao": _texto(obter(valores, "organizacao")),
                "responsavel_atendimento_texto": _texto(obter(valores, "responsavel")),
                "servidor_texto": _texto(obter(valores, "servidor")),
                "unidade": _texto(obter(valores, "unidade")),
                "quantidade_publico": publico,
                "briefing": _texto(obter(valores, "briefing")),
                "materia_site": _texto(obter(valores, "materia")),
                "origem_importacao": f"{aba}:linha {numero}",
            }
            demanda, criada = DemandaEvento.objects.update_or_create(chave_importacao=chave, defaults=defaults)
            demanda.setores.set([ascom])
            resumo["demandas" if criada else "ignoradas"] += 1
